# excel_generator.py

import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class ExcelGenerator:
    """
    Класс для создания и заполнения Excel-файла для анализа ниши.
    """

    def __init__(self, niche_name: str):
        self.workbook = Workbook()
        self.sheet = self.workbook.active
        self.sheet.title = f"Анализ - {niche_name[:20]}"

        self._setup_styles_and_headers()

    def _setup_styles_and_headers(self):
        """
        Создает шапку таблицы, объединяет ячейки и применяет стили.
        """
        self.fill_whales = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        self.fill_small = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")
        self.fill_tiny = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")

        header_font = Font(bold=True)
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))

        headers_whales = ["Киты (название канала)", "Подписчики", "Просмотры", "Идеи", "Фишки и качество"]
        for col_idx, header in enumerate(headers_whales, 1):
            cell = self.sheet.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = self.fill_whales
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
            self.sheet.column_dimensions[cell.column_letter].width = 30

        headers_small = ["Маленькие каналы", "Подписчики", "Просмотры", "Идеи", "Фишки и качество"]
        for col_idx, header in enumerate(headers_small, 7):
            cell = self.sheet.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = self.fill_small
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
            self.sheet.column_dimensions[cell.column_letter].width = 30

        headers_tiny = ["Совсем маленькие", "Подписчики", "Просмотры", "Идеи", "Фишки и качество"]
        for col_idx, header in enumerate(headers_tiny, 13):
            cell = self.sheet.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = self.fill_tiny
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
            self.sheet.column_dimensions[cell.column_letter].width = 30

        self.sheet.row_dimensions[1].height = 40

    def add_channel_data(self, category: str, data: dict):
        """
        Добавляет строку с данными о канале в нужную категорию.
        """

        start_col = 1
        if category == 'small':
            start_col = 7
        elif category == 'tiny':
            start_col = 13

        row_to_write = 2
        while self.sheet.cell(row=row_to_write, column=start_col).value is not None:
            row_to_write += 1

        # Название канала (с гиперссылкой)
        cell_name = self.sheet.cell(row=row_to_write, column=start_col)
        cell_name.value = data['name']
        cell_name.hyperlink = data['url']
        cell_name.font = Font(color="0000FF", underline="single")

        # Подписчики
        cell_subs = self.sheet.cell(row=row_to_write, column=start_col + 1)
        cell_subs.value = int(data['subs'])
        cell_subs.number_format = '#,##0'

        # Просмотры
        cell_views = self.sheet.cell(row=row_to_write, column=start_col + 2)
        cell_views.value = int(data['views'])
        cell_views.number_format = '#,##0'

        # ⬇️ --- ⭐️ ИСПРАВЛЕНИЕ ЗДЕСЬ ⭐️ --- ⬇️

        # Идеи (7, 14, 30 дней)
        cell_ideas = self.sheet.cell(row=row_to_write, column=start_col + 3)

        # Функция для создания части формулы
        def create_hyperlink_part(text, url):
            # Используем ОДИНАРНЫЕ кавычки, чтобы избежать ошибки f-string
            safe_url = str(url).replace('"', '""')
            safe_text = str(text).replace('"', '""')

            if str(url).startswith('http'):
                return f'HYPERLINK("{safe_url}", "{safe_text}")'
            else:
                return f'"{safe_text}"'

        parts = [
            create_hyperlink_part(f"7d: {data['idea_7d']}", data['idea_7d']),
            create_hyperlink_part(f"14d: {data['idea_14d']}", data['idea_14d']),
            create_hyperlink_part(f"30d: {data['idea_30d']}", data['idea_30d'])
        ]

        cell_ideas.value = f"={parts[0]} & CHAR(10) & {parts[1]} & CHAR(10) & {parts[2]}"
        cell_ideas.alignment = Alignment(wrap_text=True, horizontal='left', vertical='top')

        # ⬆️ --- ⭐️ КОНЕЦ ИСПРАВЛЕНИЯ ⭐️ --- ⬆️

        # Фишки и качество
        cell_features = self.sheet.cell(row=row_to_write, column=start_col + 4)
        cell_features.value = ""

        # Границы
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        for col_idx in range(start_col, start_col + 5):
            self.sheet.cell(row=row_to_write, column=col_idx).border = thin_border

        self.sheet.row_dimensions[row_to_write].height = 60

    def save_to_buffer(self) -> io.BytesIO:
        """
        Сохраняет Excel-книгу в буфер в памяти и возвращает его.
        """
        buffer = io.BytesIO()
        self.workbook.save(buffer)
        buffer.seek(0)
        return buffer